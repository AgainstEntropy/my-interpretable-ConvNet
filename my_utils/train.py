# -*- coding: utf-8 -*-
# @Date    : 2022/4/30 15:52
# @Author  : WangYihao
# @File    : train.py
import os
import time

import torch
import adabound
from openpyxl import load_workbook
from torch import optim, nn
from torch.utils.tensorboard import SummaryWriter
from torchvision import transforms

from my_utils import data
from my_utils.models import create_model
from my_utils.utils import get_device, check_accuracy, save_model


def trainer(model, optimizer, scheduler, loss_fn, train_loader,
            check_fn, check_loaders, batch_step, save_dir, log_every=10, epochs=2, writer=None):
    """

    Args:
        batch_step (int):
        epochs (int):
        log_every (int): log info per log_every batches.
        writer :

    Returns:
        batch_step (int):
    """
    device = get_device(model)
    # batch_size = train_loader.batch_size
    check_loader_train = check_loaders['train']
    check_loader_val = check_loaders['val']
    iters = len(train_loader)
    max_val_acc = 0.75

    for epoch in range(1, epochs + 1):
        tic = time.time()
        for batch_idx, (X, Y) in enumerate(train_loader):
            batch_step += 1
            model.train()
            X = X.to(device, dtype=torch.float32)
            Y = Y.to(device, dtype=torch.int64)
            # print(X.device, model.device)
            scores = model(X)
            loss = loss_fn(scores, Y)
            if writer is not None:
                writer.add_scalar('Metric/loss', loss.item(), batch_step)
                writer.add_scalar('Hpara/lr', optimizer.param_groups[0]['lr'], batch_step)

            # back propagate
            optimizer.zero_grad()
            loss.backward()
            optimizer.step()
            scheduler.step(batch_step / iters)

            # check accuracy
            if batch_idx % log_every == 0:
                model.eval()
                train_acc = check_fn(model, check_loader_train, training=True)
                val_acc = check_fn(model, check_loader_val, training=True)
                if writer is not None:
                    writer.add_scalars('Metric/acc', {'train': train_acc, 'val': val_acc}, batch_step)
                print(f'Epoch: {epoch} [{batch_idx}/{iters}]\tLoss: {loss:.4f}\t'
                      f'Val acc: {100. * val_acc:.1f}%')
                if val_acc > max_val_acc:
                    max_val_acc = val_acc
                    save_model(model, optimizer, scheduler,
                               save_dir=save_dir, acc=100 * val_acc)

        print(f'====> Epoch: {epoch}\tTime: {time.time() - tic}s')

    return batch_step


def train_a_model(model_configs=None, train_configs=None, loader_kwargs=None):
    """
        Train a model from zero.
    """
    if train_configs is None:
        train_configs = {
            'log_dir': 'finalruns',
            'dataset_dir': '/home/wangyh/01-Projects/03-my/Datasets/polygons_unfilled_64_3',
            'batch_size': 256,
            'epochs': 50,
            'device': 'cuda:7',
            'optim': 'Adam',
            'lr': 1e-4,
            'schedule': 'cosine_warm',
            'cos_T': 15,
            'cos_mul': 2,
            'cos_iters': 3,
            'momentum': 0.9,
            'weight_decay': 0.05,
        }
    # make dataset
    fig_resize = 64
    # mean, std = torch.tensor(0.2036), torch.tensor(0.4027)  # polygons_unfilled_32_2
    mean, std = torch.tensor(0.1094), torch.tensor(0.3660)  # polygons_unfilled_64_3
    T = transforms.Compose([
        transforms.ToTensor(),
        transforms.Resize((fig_resize, fig_resize)),
        transforms.Normalize(mean, std)
    ])
    if loader_kwargs is None:
        loader_kwargs = {
            'batch_size': train_configs['batch_size'],  # default:1
            'shuffle': True,  # default:False
            'num_workers': 4,  # default:0
            'pin_memory': True,  # default:False
            'drop_last': True,  # default:False
            'prefetch_factor': 4,  # default:2
            'persistent_workers': False  # default:False
        }
    train_loader, test_loader, check_loaders = data.make_datasets(
        dataset_dir=train_configs['dataset_dir'],
        loader_kwargs=loader_kwargs,
        transform=T
    )

    # create model
    if model_configs is None:
        model_configs = {
            'type': 'simple_conv',
            'kernel_size': 3,
            'depths': (1, 1, 1),
            'dims': (4, 8, 16)
        }
    model, optimizer, scheduler = [None] * 3
    # define model
    model = create_model(**model_configs)
    model = model.to(train_configs['device'])

    # define optimizer
    if train_configs['optim'] == 'Adam':
        optimizer = optim.Adam(params=[{'params': model.parameters(), 'initial_lr': train_configs['lr']}],
                               lr=train_configs['lr'],
                               weight_decay=train_configs['weight_decay'])
    elif train_configs['optim'] == 'AdamW':
        optimizer = optim.AdamW(model.parameters(),
                                lr=train_configs['lr'],
                                weight_decay=train_configs['weight_decay'])
    elif train_configs['optim'] == 'AdaBound':
        optimizer = adabound.AdaBound(model.parameters(),
                                      lr=train_configs['lr'],
                                      weight_decay=train_configs['weight_decay'],
                                      final_lr=0.1)
    elif train_configs['optim'] == 'SGD':
        optimizer = optim.SGD(model.parameters(),
                              lr=train_configs['lr'],
                              weight_decay=train_configs['weight_decay'],
                              momentum=train_configs['momentum'])

    # define lr scheduler
    if train_configs['schedule'] == 'cosine_warm':
        train_configs['epochs'] = int((train_configs['cos_mul'] ** train_configs['cos_iters'] - 1) / \
                                      (train_configs['cos_mul'] - 1) * train_configs['cos_T'])
        scheduler = optim.lr_scheduler.CosineAnnealingWarmRestarts(optimizer,
                                                                   T_0=train_configs['cos_T'], T_mult=2)
    elif train_configs['schedule'] == 'cosine_anneal':
        scheduler = optim.lr_scheduler.CosineAnnealingLR(optimizer, T_max=train_configs['cos_T'])
    loss_func = nn.CrossEntropyLoss()
    print(f"model ({model_configs['type']}) is on {next(model.parameters()).device}")

    # tensorboard writer
    save_time = time.strftime("%Y_%m_%d_%H_%M_%S", time.localtime())
    log_dir = os.path.join(os.getcwd(),
                           f"{train_configs['log_dir']}/{model_configs['type']}/"
                           f"KS_{model_configs['kernel_size']}_"
                           f"ACT_{model_configs['act']}_"
                           f"NM_{model_configs['norm']}")
    log_dir = os.path.join(log_dir, save_time)
    writer = SummaryWriter(log_dir=log_dir)
    with open(os.path.join(log_dir, 'para.txt'), mode='w') as f:
        f.write('## -- model configs -- ##\n')
        for k, v in model_configs.items():
            f.write(f'{k} :\t{v}\n')

        f.write('\n## -- dataset configs -- ##\n')
        for k, v in loader_kwargs.items():
            f.write(f'{k} :\t{v}\n')

        f.write('\n## -- train configs -- ##\n')
        for k, v in train_configs.items():
            f.write(f'{k} :\t{v}\n')

    # record some configs
    xlsx_path = '/home/wangyh/01-Projects/03-my/records/train_paras.xlsx'
    wb = load_workbook(xlsx_path)
    ws = wb['Sheet1']
    record_data = [model_configs['type'],
                   model_configs['kernel_size'],
                   sum(model_configs['depths']),
                   loader_kwargs['batch_size'],
                   train_configs['lr'],
                   train_configs['weight_decay'],
                   train_configs['optim'],
                   train_configs['schedule'],
                   train_configs['cos_T'],
                   train_configs['epochs']]
    ws.append(record_data)
    wb.save(filename=xlsx_path)
    row = len(list(ws.values))

    save_dir = os.path.join(log_dir, 'weights')
    os.mkdir(save_dir)
    # start training!
    trainer(model=model, optimizer=optimizer,
            scheduler=scheduler,
            loss_fn=loss_func,
            train_loader=train_loader,
            check_fn=check_accuracy,
            check_loaders=check_loaders,
            batch_step=0, epochs=train_configs['epochs'], log_every=40000 // train_configs['batch_size'] // 4,
            save_dir=save_dir,
            writer=writer)

    writer.close()

    final_acc = check_accuracy(model, test_loader, True)
    save_model(model, optimizer, scheduler, save_dir=save_dir, acc=100 * final_acc)

    ws.cell(column=len(record_data) + 1, row=row, value=final_acc)
    wb.save(filename=xlsx_path)
